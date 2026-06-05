# SPDX-License-Identifier: Apache-2.0
"""Legacy Office derivative bridge tests for the document primitive."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ummaya.tools.documents.conversion import DocumentConversionRegistry
from ummaya.tools.documents.engines import DocumentEngineRegistry
from ummaya.tools.documents.models import (
    DocumentArtifact,
    DocumentExtraction,
    DocumentFormat,
    DocumentPatch,
    FormField,
    ParagraphBlock,
    ToolResultStatus,
)
from ummaya.tools.documents.registry import DocumentToolRuntime
from ummaya.tools.documents.tool_defs import (
    DocumentFieldPatch,
    DocumentLocator,
    DocumentPrimitiveRequest,
)


def test_document_primitive_fills_legacy_doc_via_docx_derivative_bridge(
    tmp_path: Path,
) -> None:
    source = tmp_path / "legacy-application.doc"
    source.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy-doc")
    destination = tmp_path / "exports" / "legacy-application-filled.docx"
    conversion = _FakeLegacyOfficeConversionEngine(
        source_format=DocumentFormat.doc,
        output_format=DocumentFormat.docx,
        payload=b"converted-docx-payload",
    )
    conversion_registry = DocumentConversionRegistry()
    conversion_registry.register(conversion)
    engine_registry = DocumentEngineRegistry()
    engine_registry.register(_FakeLegacyDocxEngine())
    runtime = DocumentToolRuntime(
        session_id="legacy-doc-derivative",
        artifact_root=tmp_path / "store",
        engine_registry=engine_registry,
        conversion_registry=conversion_registry,
    )

    result = runtime.document(
        DocumentPrimitiveRequest(
            correlation_id="legacy-doc-derivative",
            document=DocumentLocator(path=str(source), expected_format=DocumentFormat.doc),
            operation="fill",
            instruction=f"성명 필드를 홍길동으로 작성하고 {destination}에 저장해.",
            patches=(DocumentFieldPatch(target_path="성명", value="홍길동"),),
            destination_path=str(destination),
        )
    )

    assert result.status is ToolResultStatus.ok
    assert conversion.source_artifact_id == "source-legacy-doc-derivative"
    assert result.saved_exports
    assert result.saved_exports[0].local_path == destination
    assert destination.read_bytes().endswith(b"patched:fill-001")
    assert source.read_bytes() == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy-doc"
    source_artifact = runtime.store.load_artifact("source-legacy-doc-derivative")
    working_artifact = runtime.store.load_artifact("working-legacy-doc-derivative")
    derivative_artifact = runtime.store.load_artifact("derivative-legacy-doc-derivative")
    assert source_artifact is not None
    assert working_artifact is not None
    assert derivative_artifact is not None
    assert source_artifact.format is DocumentFormat.doc
    assert working_artifact.format is DocumentFormat.docx
    assert derivative_artifact.format is DocumentFormat.docx
    assert working_artifact.parent_artifact_id == source_artifact.artifact_id
    workflow_statuses = {step.step_id: step.status.value for step in result.workflow_steps}
    assert workflow_statuses["working_copy"] == "completed"
    assert workflow_statuses["fill_style"] == "completed"
    assert workflow_statuses["diff"] == "completed"
    assert workflow_statuses["save"] == "completed"


def test_document_primitive_fills_legacy_xls_via_xlsx_derivative_bridge(
    tmp_path: Path,
) -> None:
    source = tmp_path / "legacy-application.xls"
    source.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy-xls")
    destination = tmp_path / "exports" / "legacy-application-filled.xlsx"
    conversion = _FakeLegacyOfficeConversionEngine(
        source_format=DocumentFormat.xls,
        output_format=DocumentFormat.xlsx,
        payload=_xlsx_payload(sheet_name="제출서류", cell="B1", value="13주차"),
    )
    conversion_registry = DocumentConversionRegistry()
    conversion_registry.register(conversion)
    runtime = DocumentToolRuntime(
        session_id="legacy-xls-derivative",
        artifact_root=tmp_path / "store",
        conversion_registry=conversion_registry,
    )

    result = runtime.document(
        DocumentPrimitiveRequest(
            correlation_id="legacy-xls-derivative",
            document=DocumentLocator(path=str(source), expected_format=DocumentFormat.xls),
            operation="fill",
            instruction=f"제출서류 시트 B1을 14주차로 작성하고 {destination}에 저장해.",
            patches=(DocumentFieldPatch(target_path="/sheets/제출서류/cells/B1", value="14주차"),),
            destination_path=str(destination),
        )
    )

    assert result.status is ToolResultStatus.ok
    assert conversion.source_artifact_id == "source-legacy-xls-derivative"
    assert result.saved_exports
    assert result.saved_exports[0].local_path == destination
    assert result.render_artifacts
    assert result.diff is not None
    assert load_workbook(destination)["제출서류"]["B1"].value == "14주차"
    assert source.read_bytes() == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy-xls"
    source_artifact = runtime.store.load_artifact("source-legacy-xls-derivative")
    working_artifact = runtime.store.load_artifact("working-legacy-xls-derivative")
    derivative_artifact = runtime.store.load_artifact("derivative-legacy-xls-derivative")
    assert source_artifact is not None
    assert working_artifact is not None
    assert derivative_artifact is not None
    assert source_artifact.format is DocumentFormat.xls
    assert working_artifact.format is DocumentFormat.xlsx
    assert derivative_artifact.format is DocumentFormat.xlsx
    workflow_statuses = {step.step_id: step.status.value for step in result.workflow_steps}
    assert workflow_statuses["working_copy"] == "completed"
    assert workflow_statuses["fill_style"] == "completed"
    assert workflow_statuses["diff"] == "completed"
    assert workflow_statuses["render"] == "completed"
    assert workflow_statuses["save"] == "completed"


class _FakeLegacyOfficeConversionEngine:
    engine_id = "fake-legacy-office-to-ooxml"

    def __init__(
        self,
        *,
        source_format: DocumentFormat,
        output_format: DocumentFormat,
        payload: bytes,
    ) -> None:
        self.source_format = source_format
        self.output_format = output_format
        self.payload = payload
        self.source_artifact_id: str | None = None

    def convert_for_edit(self, source: DocumentArtifact) -> bytes:
        self.source_artifact_id = source.artifact_id
        assert source.format is self.source_format
        return self.payload


def _xlsx_payload(*, sheet_name: str, cell: str, value: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet[cell] = value
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class _FakeLegacyDocxEngine:
    document_format = DocumentFormat.docx
    engine_id = "fake-legacy-docx-engine"
    render_engine_id = "fake-legacy-docx-render"
    render_artifact_extension = "txt"
    render_mime_type = "text/plain"

    def inspect(self, path: Path, *, artifact_id: str) -> DocumentExtraction:
        return DocumentExtraction(
            artifact_id=artifact_id,
            paragraphs=[
                ParagraphBlock(
                    block_id="paragraph-001",
                    text=f"Converted DOCX extracted from {path.name}",
                    source_path="/word/document.xml/p[1]",
                )
            ],
            fields=[
                FormField(
                    field_id="name",
                    label="성명",
                    path="/word/document.xml/field[name]",
                    field_type="text",
                    required=True,
                    current_value="",
                    source_confidence=Decimal("1"),
                )
            ],
            metadata={"engine_id": self.engine_id},
        )

    def apply_patch(self, path: Path, patch: DocumentPatch) -> bytes:
        payload = path.read_bytes()
        marker = "|".join(operation.operation_id for operation in patch.operations)
        return payload + f"\npatched:{marker}".encode()

    def render(self, path: Path, *, artifact_id: str, output_dir: Path) -> tuple[bytes, ...]:
        _ = output_dir
        return (f"render:{artifact_id}:{path.name}:{path.read_bytes()!r}".encode(),)
